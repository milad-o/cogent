"""
Spreadsheet capability - Excel and CSV manipulation.

Provides tools for reading, writing, and analyzing spreadsheet data,
enabling agents to work with tabular data in various formats.

Example:
    ```python
    from agenticflow import Agent
    from agenticflow.capabilities import Spreadsheet
    
    agent = Agent(
        name="DataAnalyst",
        model=model,
        capabilities=[Spreadsheet()],
    )
    
    # Agent can now work with spreadsheets
    await agent.run("Read sales.xlsx and summarize the Q4 revenue")
    await agent.run("Create a CSV report with the top 10 customers")
    ```
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from agenticflow.capabilities.base import BaseCapability
from agenticflow.tools.base import tool


@dataclass
class SheetInfo:
    """Information about a spreadsheet."""
    
    path: str
    sheets: list[str]
    active_sheet: str
    row_count: int
    column_count: int
    columns: list[str]
    
    def to_dict(self) -> dict[str, Any]:
        return {
            "path": self.path,
            "sheets": self.sheets,
            "active_sheet": self.active_sheet,
            "row_count": self.row_count,
            "column_count": self.column_count,
            "columns": self.columns,
        }


@dataclass
class ReadResult:
    """Result of reading spreadsheet data."""
    
    data: list[dict[str, Any]]
    columns: list[str]
    row_count: int
    sheet_name: str | None = None
    
    def to_dict(self) -> dict[str, Any]:
        # Limit data for display
        display_data = self.data[:100] if len(self.data) > 100 else self.data
        return {
            "data": display_data,
            "columns": self.columns,
            "row_count": self.row_count,
            "sheet_name": self.sheet_name,
            "truncated": len(self.data) > 100,
        }


@dataclass
class WriteResult:
    """Result of writing spreadsheet data."""
    
    path: str
    rows_written: int
    sheet_name: str | None = None
    success: bool = True
    error: str | None = None
    
    def to_dict(self) -> dict[str, Any]:
        return {
            "path": self.path,
            "rows_written": self.rows_written,
            "sheet_name": self.sheet_name,
            "success": self.success,
            "error": self.error,
        }


class Spreadsheet(BaseCapability):
    """
    Spreadsheet capability for Excel and CSV manipulation.
    
    Provides tools for reading, writing, and analyzing tabular data
    in Excel (.xlsx, .xls) and CSV formats.
    
    Args:
        allowed_paths: List of paths the agent can access. If empty, allows
            current working directory only.
        max_rows: Maximum rows to read at once (default: 100000)
        max_file_size_mb: Maximum file size to process in MB (default: 50)
        
    Tools provided:
        - read_spreadsheet: Read data from Excel/CSV file
        - write_spreadsheet: Write data to Excel/CSV file
        - get_sheet_info: Get metadata about a spreadsheet
        - query_spreadsheet: Query data with filters
        - aggregate_spreadsheet: Compute aggregations (sum, avg, count, etc.)
        - merge_spreadsheets: Merge multiple spreadsheets
    """
    
    @property
    def name(self) -> str:
        """Unique name for this capability."""
        return "spreadsheet"
    
    @property
    def tools(self) -> list:
        """Tools this capability provides to the agent."""
        return self.get_tools()
    
    def __init__(
        self,
        allowed_paths: list[str | Path] | None = None,
        max_rows: int = 100000,
        max_file_size_mb: int = 50,
    ) -> None:
        self.allowed_paths = [Path(p).resolve() for p in (allowed_paths or ["."])]
        self.max_rows = max_rows
        self.max_file_size_mb = max_file_size_mb
        
        # Check for optional dependencies
        import importlib.util
        self._has_openpyxl = importlib.util.find_spec("openpyxl") is not None
        self._has_pandas = importlib.util.find_spec("pandas") is not None
    
    def _validate_path(self, path: str | Path) -> Path:
        """Validate that path is within allowed directories."""
        path = Path(path).resolve()
        
        for allowed in self.allowed_paths:
            try:
                path.relative_to(allowed)
                return path
            except ValueError:
                continue
        
        msg = f"Path {path} is not within allowed directories"
        raise PermissionError(msg)
    
    def _check_file_size(self, path: Path) -> None:
        """Check file size is within limits."""
        if path.exists():
            size_mb = path.stat().st_size / (1024 * 1024)
            if size_mb > self.max_file_size_mb:
                msg = f"File size ({size_mb:.1f}MB) exceeds limit ({self.max_file_size_mb}MB)"
                raise ValueError(msg)
    
    def _detect_format(self, path: Path) -> str:
        """Detect file format from extension."""
        suffix = path.suffix.lower()
        if suffix in (".xlsx", ".xlsm", ".xltx", ".xltm"):
            return "xlsx"
        elif suffix in (".xls",):
            return "xls"
        elif suffix in (".csv",):
            return "csv"
        elif suffix in (".tsv",):
            return "tsv"
        else:
            # Try to detect from content
            return "csv"
    
    # =========================================================================
    # CSV Operations (no dependencies)
    # =========================================================================
    
    def _read_csv(
        self,
        path: Path,
        delimiter: str = ",",
        has_header: bool = True,
        max_rows: int | None = None,
    ) -> ReadResult:
        """Read CSV file."""
        data = []
        columns = []
        
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f, delimiter=delimiter)
            
            if has_header:
                columns = next(reader, [])
            
            max_rows = max_rows or self.max_rows
            for i, row in enumerate(reader):
                if i >= max_rows:
                    break
                
                if has_header:
                    data.append(dict(zip(columns, row)))
                else:
                    if not columns:
                        columns = [f"col_{j}" for j in range(len(row))]
                    data.append(dict(zip(columns, row)))
        
        return ReadResult(
            data=data,
            columns=columns,
            row_count=len(data),
        )
    
    def _write_csv(
        self,
        path: Path,
        data: list[dict[str, Any]],
        delimiter: str = ",",
    ) -> WriteResult:
        """Write data to CSV file."""
        if not data:
            return WriteResult(path=str(path), rows_written=0)
        
        columns = list(data[0].keys())
        
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=columns, delimiter=delimiter)
            writer.writeheader()
            writer.writerows(data)
        
        return WriteResult(
            path=str(path),
            rows_written=len(data),
        )
    
    # =========================================================================
    # Excel Operations (requires openpyxl)
    # =========================================================================
    
    def _read_excel(
        self,
        path: Path,
        sheet_name: str | None = None,
        has_header: bool = True,
        max_rows: int | None = None,
    ) -> ReadResult:
        """Read Excel file using openpyxl."""
        if not self._has_openpyxl:
            msg = "openpyxl not installed. Install with: pip install openpyxl"
            raise ImportError(msg)
        
        from openpyxl import load_workbook
        
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        data = []
        columns = []
        max_rows = max_rows or self.max_rows
        
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 and has_header:
                columns = [str(c) if c else f"col_{j}" for j, c in enumerate(row)]
                continue
            
            if i >= max_rows + (1 if has_header else 0):
                break
            
            if not columns:
                columns = [f"col_{j}" for j in range(len(row))]
            
            row_dict = {}
            for j, value in enumerate(row):
                col_name = columns[j] if j < len(columns) else f"col_{j}"
                row_dict[col_name] = value
            data.append(row_dict)
        
        wb.close()
        
        return ReadResult(
            data=data,
            columns=columns,
            row_count=len(data),
            sheet_name=ws.title,
        )
    
    def _write_excel(
        self,
        path: Path,
        data: list[dict[str, Any]],
        sheet_name: str = "Sheet1",
    ) -> WriteResult:
        """Write data to Excel file using openpyxl."""
        if not self._has_openpyxl:
            msg = "openpyxl not installed. Install with: pip install openpyxl"
            raise ImportError(msg)
        
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        if not data:
            wb.save(path)
            return WriteResult(path=str(path), rows_written=0, sheet_name=sheet_name)
        
        columns = list(data[0].keys())
        
        # Write header
        for j, col in enumerate(columns, 1):
            ws.cell(row=1, column=j, value=col)
        
        # Write data
        for i, row in enumerate(data, 2):
            for j, col in enumerate(columns, 1):
                ws.cell(row=i, column=j, value=row.get(col))
        
        wb.save(path)
        
        return WriteResult(
            path=str(path),
            rows_written=len(data),
            sheet_name=sheet_name,
        )
    
    def _get_excel_info(self, path: Path) -> SheetInfo:
        """Get information about Excel file."""
        if not self._has_openpyxl:
            msg = "openpyxl not installed. Install with: pip install openpyxl"
            raise ImportError(msg)
        
        from openpyxl import load_workbook
        
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        
        # Get columns from first row
        columns = []
        for cell in next(ws.iter_rows(max_row=1, values_only=True), []):
            columns.append(str(cell) if cell else "")
        
        info = SheetInfo(
            path=str(path),
            sheets=wb.sheetnames,
            active_sheet=ws.title,
            row_count=ws.max_row - 1 if ws.max_row else 0,  # Exclude header
            column_count=ws.max_column or 0,
            columns=columns,
        )
        
        wb.close()
        return info
    
    # =========================================================================
    # Data Operations
    # =========================================================================
    
    def _filter_data(
        self,
        data: list[dict[str, Any]],
        filters: dict[str, Any],
    ) -> list[dict[str, Any]]:
        """Filter data based on conditions.
        
        Supports:
            - Exact match: {"column": "value"}
            - Comparison: {"column": {"$gt": 10, "$lt": 100}}
            - Contains: {"column": {"$contains": "text"}}
            - In list: {"column": {"$in": [1, 2, 3]}}
        """
        result = []
        
        for row in data:
            match = True
            for col, condition in filters.items():
                value = row.get(col)
                
                if isinstance(condition, dict):
                    # Complex condition
                    for op, target in condition.items():
                        if op == "$gt" and not (value is not None and value > target):
                            match = False
                        elif op == "$gte" and not (value is not None and value >= target):
                            match = False
                        elif op == "$lt" and not (value is not None and value < target):
                            match = False
                        elif op == "$lte" and not (value is not None and value <= target):
                            match = False
                        elif op == "$ne" and value == target:
                            match = False
                        elif op == "$in" and value not in target:
                            match = False
                        elif op == "$nin" and value in target:
                            match = False
                        elif op == "$contains":
                            if not (value and str(target).lower() in str(value).lower()):
                                match = False
                        elif op == "$startswith":
                            if not (value and str(value).lower().startswith(str(target).lower())):
                                match = False
                        elif op == "$endswith":
                            if not (value and str(value).lower().endswith(str(target).lower())):
                                match = False
                else:
                    # Exact match
                    if value != condition:
                        match = False
                
                if not match:
                    break
            
            if match:
                result.append(row)
        
        return result
    
    def _aggregate_data(
        self,
        data: list[dict[str, Any]],
        group_by: list[str] | None,
        aggregations: dict[str, str],
    ) -> list[dict[str, Any]]:
        """Aggregate data with grouping.
        
        Aggregation functions: sum, avg, min, max, count, first, last
        """
        if not data:
            return []
        
        # Group data
        groups: dict[tuple, list[dict]] = {}
        
        for row in data:
            if group_by:
                key = tuple(row.get(col) for col in group_by)
            else:
                key = ("__all__",)
            
            if key not in groups:
                groups[key] = []
            groups[key].append(row)
        
        # Compute aggregations
        results = []
        for key, rows in groups.items():
            result_row = {}
            
            # Add group columns
            if group_by:
                for i, col in enumerate(group_by):
                    result_row[col] = key[i]
            
            # Compute aggregations
            for col, func in aggregations.items():
                values = [r.get(col) for r in rows if r.get(col) is not None]
                numeric = [v for v in values if isinstance(v, (int, float))]
                
                if func == "sum":
                    result_row[f"{col}_sum"] = sum(numeric) if numeric else 0
                elif func == "avg":
                    result_row[f"{col}_avg"] = sum(numeric) / len(numeric) if numeric else 0
                elif func == "min":
                    result_row[f"{col}_min"] = min(numeric) if numeric else None
                elif func == "max":
                    result_row[f"{col}_max"] = max(numeric) if numeric else None
                elif func == "count":
                    result_row[f"{col}_count"] = len(values)
                elif func == "first":
                    result_row[f"{col}_first"] = values[0] if values else None
                elif func == "last":
                    result_row[f"{col}_last"] = values[-1] if values else None
            
            results.append(result_row)
        
        return results
    
    # =========================================================================
    # Tool Methods
    # =========================================================================
    
    def get_tools(self) -> list:
        """Return the tools provided by this capability."""
        
        @tool
        def read_spreadsheet(
            path: str,
            sheet_name: str | None = None,
            has_header: bool = True,
            max_rows: int = 1000,
        ) -> dict[str, Any]:
            """Read data from an Excel or CSV file.
            
            Args:
                path: Path to the spreadsheet file (.xlsx, .csv, .tsv)
                sheet_name: Sheet name for Excel files (uses active sheet if not specified)
                has_header: Whether first row contains column headers
                max_rows: Maximum rows to read (default: 1000)
            
            Returns:
                Dictionary with data, columns, row_count, and sheet_name
            """
            file_path = self._validate_path(path)
            self._check_file_size(file_path)
            
            # Handle string conversion
            if isinstance(max_rows, str):
                max_rows = int(max_rows) if max_rows else 1000
            
            fmt = self._detect_format(file_path)
            
            if fmt == "xlsx":
                result = self._read_excel(file_path, sheet_name, has_header, max_rows)
            elif fmt == "csv":
                result = self._read_csv(file_path, ",", has_header, max_rows)
            elif fmt == "tsv":
                result = self._read_csv(file_path, "\t", has_header, max_rows)
            else:
                result = self._read_csv(file_path, ",", has_header, max_rows)
            
            return result.to_dict()
        
        @tool
        def write_spreadsheet(
            path: str,
            data: list[dict[str, Any]],
            sheet_name: str = "Sheet1",
        ) -> dict[str, Any]:
            """Write data to an Excel or CSV file.
            
            Args:
                path: Path for the output file (.xlsx or .csv)
                data: List of dictionaries (rows), each dict is column_name: value
                sheet_name: Sheet name for Excel files
            
            Returns:
                Dictionary with path, rows_written, and success status
            """
            file_path = self._validate_path(path)
            fmt = self._detect_format(file_path)
            
            if fmt == "xlsx":
                result = self._write_excel(file_path, data, sheet_name)
            elif fmt == "tsv":
                result = self._write_csv(file_path, data, "\t")
            else:
                result = self._write_csv(file_path, data, ",")
            
            return result.to_dict()
        
        @tool
        def get_sheet_info(path: str) -> dict[str, Any]:
            """Get information about a spreadsheet file.
            
            Args:
                path: Path to the spreadsheet file
            
            Returns:
                Dictionary with sheets, columns, row_count, etc.
            """
            file_path = self._validate_path(path)
            fmt = self._detect_format(file_path)
            
            if fmt == "xlsx":
                return self._get_excel_info(file_path).to_dict()
            else:
                # CSV info
                result = self._read_csv(file_path, max_rows=1)
                return {
                    "path": str(file_path),
                    "format": "csv",
                    "columns": result.columns,
                    "sample_row": result.data[0] if result.data else {},
                }
        
        @tool
        def query_spreadsheet(
            path: str,
            columns: str | None = None,
            filter_column: str | None = None,
            filter_value: str | None = None,
            sort_by: str | None = None,
            sort_desc: bool = False,
            limit: int = 100,
            sheet_name: str | None = None,
        ) -> dict[str, Any]:
            """Query and filter spreadsheet data.
            
            Args:
                path: Path to the spreadsheet file
                columns: Comma-separated column names to include (e.g., "name,sales,region")
                filter_column: Column name to filter on
                filter_value: Value to filter for (exact match)
                sort_by: Column name to sort by
                sort_desc: Sort in descending order
                limit: Maximum rows to return
                sheet_name: Sheet name for Excel files
            
            Returns:
                Filtered and sorted data
            """
            file_path = self._validate_path(path)
            self._check_file_size(file_path)
            
            # Handle string limit
            if isinstance(limit, str):
                limit = int(limit) if limit else 100
            
            fmt = self._detect_format(file_path)
            
            if fmt == "xlsx":
                result = self._read_excel(file_path, sheet_name)
            else:
                result = self._read_csv(file_path)
            
            data = result.data
            
            # Apply simple filter
            if filter_column and filter_value:
                data = [row for row in data if str(row.get(filter_column, "")) == str(filter_value)]
            
            # Parse columns string
            col_list = None
            if columns:
                col_list = [c.strip() for c in columns.split(",")]
            
            # Select columns
            if col_list:
                data = [{k: v for k, v in row.items() if k in col_list} for row in data]
            
            # Sort
            if sort_by and data:
                data = sorted(
                    data,
                    key=lambda x: (x.get(sort_by) is None, x.get(sort_by)),
                    reverse=sort_desc,
                )
            
            # Limit
            data = data[:limit]
            
            return {
                "data": data,
                "total_matching": len(data),
                "columns": col_list or result.columns,
            }
        
        @tool
        def aggregate_spreadsheet(
            path: str,
            column: str,
            operation: str = "sum",
            group_by: str | None = None,
            sheet_name: str | None = None,
        ) -> dict[str, Any]:
            """Compute aggregation on a spreadsheet column.
            
            Args:
                path: Path to the spreadsheet file
                column: Column name to aggregate (e.g., "sales")
                operation: Aggregation function: sum, avg, min, max, count
                group_by: Column to group by (e.g., "region")
                sheet_name: Sheet name for Excel files
            
            Returns:
                Aggregated data with results
            """
            file_path = self._validate_path(path)
            self._check_file_size(file_path)
            
            fmt = self._detect_format(file_path)
            
            if fmt == "xlsx":
                result = self._read_excel(file_path, sheet_name)
            else:
                result = self._read_csv(file_path)
            
            data = result.data
            
            # Convert string numbers to floats for aggregation
            for row in data:
                if column in row and isinstance(row[column], str):
                    try:
                        row[column] = float(row[column].replace(",", ""))
                    except (ValueError, AttributeError):
                        pass
            
            # Build aggregations dict
            aggregations = {column: operation}
            group_by_list = [group_by] if group_by else None
            
            # Aggregate
            aggregated = self._aggregate_data(data, group_by_list, aggregations)
            
            return {
                "data": aggregated,
                "group_by": group_by,
                "column": column,
                "operation": operation,
                "source_rows": len(data),
            }
        
        @tool
        def merge_spreadsheets(
            paths: list[str],
            output_path: str,
            merge_type: str = "concat",
            join_column: str | None = None,
        ) -> dict[str, Any]:
            """Merge multiple spreadsheets into one.
            
            Args:
                paths: List of spreadsheet file paths to merge
                output_path: Path for the merged output file
                merge_type: "concat" (stack rows) or "join" (merge on column)
                join_column: Column to join on (required if merge_type is "join")
            
            Returns:
                Information about the merged file
            """
            if not paths:
                return {"error": "No paths provided"}
            
            all_data = []
            all_columns = set()
            
            for path in paths:
                file_path = self._validate_path(path)
                self._check_file_size(file_path)
                
                fmt = self._detect_format(file_path)
                if fmt == "xlsx":
                    result = self._read_excel(file_path)
                else:
                    result = self._read_csv(file_path)
                
                all_data.append(result.data)
                all_columns.update(result.columns)
            
            if merge_type == "concat":
                # Stack all rows
                merged = []
                for data in all_data:
                    merged.extend(data)
            elif merge_type == "join":
                if not join_column:
                    return {"error": "join_column required for join merge_type"}
                
                # Simple left join
                merged = list(all_data[0])
                for data in all_data[1:]:
                    # Build lookup
                    lookup = {row.get(join_column): row for row in data}
                    for row in merged:
                        key = row.get(join_column)
                        if key in lookup:
                            row.update(lookup[key])
            else:
                return {"error": f"Unknown merge_type: {merge_type}"}
            
            # Write output
            output_file = self._validate_path(output_path)
            fmt = self._detect_format(output_file)
            
            if fmt == "xlsx":
                write_result = self._write_excel(output_file, merged)
            else:
                write_result = self._write_csv(output_file, merged)
            
            return {
                "output_path": str(output_file),
                "rows_merged": write_result.rows_written,
                "columns": list(all_columns),
                "files_merged": len(paths),
            }
        
        return [
            read_spreadsheet,
            write_spreadsheet,
            get_sheet_info,
            query_spreadsheet,
            aggregate_spreadsheet,
            merge_spreadsheets,
        ]
