"""Excel file loader."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from agenticflow.document.loaders.base import BaseLoader
from agenticflow.document.types import Document


class XLSXLoader(BaseLoader):
    """Loader for Excel files.

    Supports .xlsx and .xls files. Each sheet becomes a separate Document.
    Requires openpyxl for .xlsx files.

    Example:
        >>> loader = XLSXLoader()
        >>> docs = await loader.load(Path("spreadsheet.xlsx"))
        >>> for doc in docs:
        ...     print(f"Sheet: {doc.metadata['sheet']}")
    """

    supported_extensions = [".xlsx", ".xls"]

    def __init__(
        self,
        encoding: str = "utf-8",
        sheet_names: list[str] | None = None,
    ) -> None:
        """Initialize the loader.

        Args:
            encoding: Not used for Excel (binary format).
            sheet_names: Specific sheets to load (None = all sheets).
        """
        super().__init__(encoding)
        self.sheet_names = sheet_names

    async def load(self, path: str | Path, **kwargs: Any) -> list[Document]:
        """Load an Excel file.

        Args:
            path: Path to the Excel file (str or Path).
            **kwargs: Additional options.

        Returns:
            List of Documents, one per sheet.

        Raises:
            ImportError: If openpyxl is not installed.
        """
        path = Path(path)
        if path.suffix.lower() == ".xls":
            raise ValueError(
                "Old .xls format requires xlrd. "
                "Please convert to .xlsx format or install xlrd."
            )

        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError(
                "Excel loading requires 'openpyxl'. "
                "Install with: uv add openpyxl"
            )

        wb = load_workbook(str(path), read_only=True, data_only=True)
        documents = []

        # Filter sheets if specified
        sheet_names = self.sheet_names or wb.sheetnames

        for sheet_name in sheet_names:
            if sheet_name not in wb.sheetnames:
                continue

            sheet = wb[sheet_name]
            rows = []
            headers: list[str] = []

            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:
                    # First row as headers
                    headers = [str(c) if c else f"col_{j}" for j, c in enumerate(row)]
                    continue

                # Skip empty rows
                if not any(c is not None for c in row):
                    continue

                row_text = ", ".join(
                    f"{headers[j]}: {c}"
                    for j, c in enumerate(row)
                    if c is not None and j < len(headers)
                )
                rows.append(row_text)

            if rows:
                documents.append(
                    self._create_document(
                        "\n".join(rows),
                        path,
                        sheet=sheet_name,
                        columns=headers,
                        row_count=len(rows),
                    )
                )

        wb.close()
        return documents


__all__ = ["XLSXLoader"]
